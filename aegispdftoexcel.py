import tkinter as tk
from tkinter import filedialog, messagebox, font, ttk

from PyPDF2 import PdfReader
import re
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Side
from datetime import datetime
import json

import os
from os.path import basename


def to_digit(data):
    for row_index, row in enumerate(data):
        for col_index, value in enumerate(row):
            if not isinstance(value, str):
                continue

            if value.isdigit():
                data[row_index][col_index] = int(value)
            else:
                try:
                    data[row_index][col_index] = float(value)
                except ValueError:
                    pass

    return data

def all_border(sheet):
    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
        for cell in row:
            cell.border = Border(
                left=Side(border_style="thin", color="000000"),
                right=Side(border_style="thin", color="000000"),
                top=Side(border_style="thin", color="000000"),
                bottom=Side(border_style="thin", color="000000")
            )

def replace_text(json_file, text):
    with open(json_file, "r") as file:
        data_dict = json.load(file)

    for key, value in data_dict.items():
        text = text.replace(key, value)

    return text

def extract_pdf_data(input_pdf_file):
    pdf_reader = PdfReader(input_pdf_file)
    all_page_details = [pdf_reader.pages[page_number].extract_text() for page_number in range(len(pdf_reader.pages))]
    return ''.join(all_page_details)

def process_text(page_details):
    default_date = page_details.split("WEEK ENDING", 1)[-1].split("ADVICE", 1)[0].strip()
    default_date = datetime.strptime(default_date, "%m/%d/%Y").date()

    contract_num = page_details.split("ISP AGREEMENT ID: ", 1)[-1].split("ISP SIGNATORY FEDEX ID:", 1)[0].strip()

    isp_id = page_details.split("ISP SIGNATORY FEDEX ID: ", 1)[-1].split("ADDRESS:", 1)[0].strip()

    station = page_details.split("PRIMARY STATION #: ", 1)[-1].split('\n')[0].strip().replace("   ", " ")

    page_details = re.split(r"CONTRACTED SERVICE AREA:[^\n]*", page_details, maxsplit=1)[-1]

    page_details = page_details.replace("_", "").replace("# ", "#").replace("$ ", "$").replace("$TOTAL AMT", "$TOTAL_AMT").replace("WEEKLY TOTALS", "WEEKLY_TOTALS").replace(" TOTALS", "").replace(" %", "%")
    page_details = re.sub(re.compile(r'Page \d{1}'), '', page_details)

    lines = page_details.split('\n')
    patterns = [
        re.compile(r'^WEEKLY INDEPENDENT.*'),
        re.compile(r'^WEEK ENDING.*'),
        re.compile(r'^ISP AGREEMENT*'),
        re.compile(r'^ENTITY ID.*'),
        re.compile(r'^ENTITY NAME.*')
    ]
    filtered_lines = [line for line in lines if not any(pattern.match(line) for pattern in patterns)]
    page_details = '\n'.join(filtered_lines)

    processed_details = re.compile(r"ISP AGREEMENT[^\n]*").sub("", page_details)

    return default_date, processed_details, isp_id, station, contract_num

def create_dataframe(default_date, processed_details, default_station, contract_num):
    dataframe_dict = {}

    # Table 1
    table1 = re.split(r"-*\s*LARGE PKG MIX\s*-+\s*-*\s*E-COMMERCE\s*-+", processed_details, maxsplit=1)[0]
    table1 = table1.split("TOTAL FUEL SURGE STOP GRAND")[1]

    # Create DataFrame1
    data1 = [line.split() for line in table1.split('\n') if line.strip()]
    columns = data1[0]
    columns.insert(0, "DEFAULT DATE")
    # rows = data1[1:]
    for inner_list in data1[1:]:
        if inner_list[0] != 'WEEKLY_TOTALS:':
            inner_list[0] = datetime.strptime(inner_list[0], '%m/%d/%y').date()

    for inner_list in data1[1:]:
        inner_list.insert(0, default_date)
        inner_list.append(contract_num)

    columns.append("Contract #")

    for i in range(1, len(data1)):
        if len(data1[i]) != 15:
            data1[i][0] = str(data1[i][0])
            data1[i][1] = str(data1[i][1])
            joined_string = " <-> ".join(data1[i])
            data1[i] = [joined_string, "", "", "", "", "", "", "", "", "", "", "", "", "", ""]
            data1[i].append("Yes")
        else:
            data1[i].append("No")
    rows = data1[1:]
    columns.append("Need Edit")
    df1 = pd.DataFrame(rows, columns=columns)


    # Table 2
    table2 = re.split(r"-*\s*LARGE PKG MIX\s*-+\s*-*\s*E-COMMERCE\s*-+", processed_details, maxsplit=1)[-1]
    table2 = re.split(r"\*eCommerce[^\n]*", table2)[0]
    lines = table2.split('\n')
    non_empty_lines = [line for line in lines if line.strip() != '']
    table2 = '\n'.join(non_empty_lines)
    lines = table2.split('\n')
    lines[0] += " $TOTAL_AMT"
    table2 = '\n'.join(lines)

    # Create DataFrame2
    data2 = [line.split() for line in table2.split('\n') if line.strip()]
    columns = data2[0]
    columns.insert(0, "DEFAULT DATE")
    # rows = data2[1:]
    for inner_list in data2[1:]:
        if inner_list[0] != 'WEEKLY_TOTALS:':
            inner_list[0] = datetime.strptime(inner_list[0], '%m/%d/%y').date()
        else:
            inner_list.insert(1, "")

    for inner_list in data2[1:]:
        inner_list.insert(0, default_date)

    for i in range(1, len(data2)):
        if len(data2[i]) != 9:
            data2[i][0] = str(data2[i][0])
            data2[i][1] = str(data2[i][1])
            joined_string = " <-> ".join(data2[i])
            data2[i] = [joined_string, "", "", "", "", "", "", "", ""]
            data2[i].append("Yes")
        else:
            data2[i].append("No")
    rows = data2[1:]
    columns.append("Need Edit")
    df2 = pd.DataFrame(rows, columns=columns)  


    # Table 3
    table3 = processed_details.split("OTHER P&D CHARGES", 1)[-1]
    table3 = table3.split("OTHER P&D CHARGES",)[0]
    lines = table3.split('\n')
    non_empty_lines = [line for line in lines if line.strip() != '']
    table3 = '\n'.join(non_empty_lines)
    table3 = table3.replace(" - ", "-").replace("  ", "").replace("Large Package Mix", "LargePackageMix").replace(" Promotion", "Promotion").replace(" Charge", "Charge").replace(" Package", "Package").replace(" Stop", "Stop").replace(" Surcharge", "Surcharge").replace(": ", ":").replace(" charge ", "charge_").replace(" trans ", "trans").replace("Safe Operating Incentive", "SafeOperatingIncentive").replace(" Q1 ", "Q1").replace(" Q2 ", "Q2").replace(" Q3 ", "Q3").replace(" Q4 ", "Q4").replace(" Contingency", "Contingency").replace(" adjust", "adjust").replace(" previous ", "previous").replace("Blind Spot ", "BlindSpot").replace(" Sensor", "Sensor").replace(" Theft ", "Theft").replace(" Time ", "Time").replace(" Activity", "Activity")

    table3 = replace_text("sheet3.json", table3)

    pattern = r'\d{1,2}/\d{1,2}/\d{4}\s\d{1,2}/\d{1,2}/\d{2}-\d{1,2}/\d{1,2}/\d{2}'
    def replace_space(match):
        return match.group().replace(' ', '|')
    table3 = re.sub(pattern, replace_space, table3)

    pattern = r'(\d{2}/\d{2}/\d{4}) ([^.]*[a-zA-Z]+)'
    table3 = re.sub(pattern, replace_space, table3)

    # Create DataFrame3
    data3 = [line.split() for line in table3.split('\n') if line.strip()]

    data3[0].insert(0, "DATE")

    for i in range(1, len(data3)):
        if len(data3[i]) == 2:
            data3[i].insert(1, "")
        data3[i].insert(0, default_date)

    for i in range(1, len(data3)):
        if len(data3[i]) == 5:
            if "Contingency" in data3[i][2] or "through" in data3[i][2]:
                joined_item = f"{data3[i][2]}|{data3[i][3]}"
                data3[i] = data3[i][:2] + [joined_item] + data3[i][4:]

    for i in range(1, len(data3)):
        if len(data3[i]) != 4:
            data3[i][0] = str(data3[i][0])
            joined_string = " <-> ".join(data3[i])
            data3[i] = [joined_string, "", "", ""]
            data3[i].append("Yes")
        else:
            data3[i].append("No")


    columns = data3[0]
    columns.append("Need Edit")
    rows = data3[1:]
    # for item in rows:
    #     print(f"length: {len(item)} | item: {item}")
    df3 = pd.DataFrame(rows, columns=columns)


    # Table 4
    if "CHARGEBACKS AND DEDUCTIONS" in processed_details:
        table4 = processed_details.split("CHARGEBACKS AND DEDUCTIONS", 1)[-1]
        table4 = table4.split("CHARGEBACKS AND DEDUCTIONS TOTAL:")[0]
        table4 = re.sub(re.compile(r"DEDUCTION DEDUCTION[^\n]*"), "", table4) 

        pattern = r'Q[1-4]\s\d{4}\s[A-Za-z]\d+'
        matches = re.finditer(pattern, table4)
        for match in matches:
            match_str = match.group()
            replaced_str = re.sub(r'\s', '-', match_str)
            table4 = table4.replace(match_str, replaced_str, 1)

        table4 = table4.replace("DOCUMENT ID", "DOCUMENTID").replace("GOAL $AMT", "DEDUCTION_GOAL$AMT").replace("$TAKEN TO DATE", "DEDUCTION_$TAKEN_TO_DATE").replace("$AMOUNT", "DEDUCTION_$AMOUNT", 1).replace(" $AMOUNT ", " SALESTAX_$AMOUNT ", 1).replace("$ARREARS", "DEDUCTION_$ARREARS").replace("$REFUND", "DEDUCTION_$REFUND").replace("NET $AMT", "DEDUCTION_NET$AMT").replace("Invoice diverted chrg stmt", "Invoice_diverted_chrg_stmt").replace("SUB TOTAL", "SUBTOTAL").replace("-", "").replace("Liability Cost Contribution", "Liability_Cost_Contribution").replace("Q1 ", "Q1-").replace("Q2 ", "Q2-").replace("Q3 ", "Q3-").replace("Q4 ", "Q4-").replace("Inv divert to Installment Plan", "Inv_divert_to_Installment_Plan")
        
        

        # Create DataFrame4
        data4 = [line.split() for line in table4.split('\n') if line.strip()]
        # for item in data4:
        #     print(f"length: {len(item)} | data_item: {item}")

        for i in range(1, 8):
            data4[1].insert(i, "")

        for j in range(2, len(data4) - 1):
            if data4[j][0] in ["Invoice_diverted_chrg_stmt", "Liability_Cost_Contribution", "Inv_divert_to_Installment_Plan"]:
                for i in range(1, 8):
                    data4[j].insert(i, "")
            else:
                data4[j].insert(0, "")
                for i in range(5, 9):
                    data4[j].insert(i, "")

        for i in range(1, 8):
            data4[-1].insert(i, "")

        data4[0].insert(0, "DATE")
        for i in range(1, len(data4)):
            data4[i].insert(0, default_date)

        # for item in rows:
        #     print(f"length: {len(item)} | item: {item}")
        for i in range(1, len(data4)):
            if len(data4[i]) != 10:
                data4[i][0] = str(data4[i][0])
                joined_string = " <-> ".join(data4[i])
                data4[i] = [joined_string, "", "", "", "", "", "", "", "", ""]
                data4[i].append("Yes")
            else:
                data4[i].append("No")
                
        columns = data4[0]
        rows = data4[1:]

        columns.append("Need Edit")
        df4 = pd.DataFrame(rows, columns=columns)
    else:
        data4 = ""
        df4 = pd.DataFrame(columns=['DATE', 'TYPE', 'DOCUMENTID', 'DEDUCTION_GOAL$AMT', 'DEDUCTION_$TAKEN_TO_DATE', 'DEDUCTION_$AMOUNT', 'SALESTAX_$AMOUNT', 'DEDUCTION_$ARREARS', 'DEDUCTION_$REFUND', 'DEDUCTION_NET$AMT', 'Need Edit'])


    # Table 5
    table5 = processed_details.split("YEAR TO DATE CHARGES", 1)[-1]
    table5 = table5.split("SETTLEMENT AMOUNT", 1)[0]
    table5 = table5.replace("BEGINNING BALANCE (YTD):", "BEGINNINGBALANCE(YTD):").replace("NEW ACTIVITY:", "NEWACTIVITY:").replace("CURRENT BALANCE:", "CURRENTBALANCE:")

    # Create DataFrame5
    data5 = [line.split() for line in table5.split('\n') if line.strip()]

    for i in range(len(data5)):
        data5[i].insert(0, default_date)

    for i in range(len(data5)):
        if len(data5[i]) != 3:
            data5[i][0] = str(data5[i][0])
            joined_string = " <-> ".join(data5[i])
            data5[i] = [joined_string, "", ""]
            data5[i].append("Yes")
        else:
            data5[i].append("No")
                
    columns = data5[0]
    rows = data5[1:]

    rows = data5
    df5 = pd.DataFrame(rows, columns=None)


    # Table 6
    table6 = re.split(r"DRIVER DRIVER[^\n]*", processed_details)[-1]
    station = re.split(r"DRIVER DRIVER[^\n]*", processed_details)[0].strip()
    station = station.split("FACILITY #:")[-1]

    if "NON-SCAN ACTIVITY" in table6:
        table6 = table6.split("NON-SCAN ACTIVITY", 1)[0]
    elif "DOUBLE STOPS:" in table6:
        table6 = table6.split("DOUBLE STOPS:", 1)[0]
    else:
        table6 = table6.split("OTHER INFORMATION:", 1)[0]

    table6 = table6.replace("FEDEX ID", "FEDEXID").replace(".", "").replace("FEDEX NAME", "FEDEXNAME").replace("-", "xx").replace("’", "xxx").replace("'", "xxxx")

    # Create DataFrame6
    data6 = [line.split() for line in table6.split('\n') if line.strip()]

    for j in range(1, len(data6)):
        merged_list = []
        i = 0
        while i < len(data6[j]):
            current_item = data6[j][i]
            merged_item = current_item
            while i < len(data6[j]) - 1 and current_item.isalpha() and data6[j][i + 1].isalpha(): 
                # Continue merging consecutive alpha items with "_"
                merged_item += "_" + data6[j][i + 1]
                i += 1 
            merged_list.append(merged_item)
            i += 1

        data6[j] = merged_list

    for i in range(1, len(data6)):
        if "WEEKLY_TOTALS:" in data6[i]:
            data6[i].insert(1, "")
            data6[i].insert(2, "")

        for j in range(len(data6[i])):
            data6[i][j] = data6[i][j].replace("_", " ").replace("xx", "-").replace("xxx", "’").replace("xxxx", "'")


    # station = default_station
    for i in range(1, len(data6)):
        if "FACILITY" in data6[i]:
            data6[i] = data6[i][2:]
            station = ' '.join(data6[i])
        elif "WEEKLY_TOTALS:" in data6[i]:
            pass
        else:
            data6[i].insert(9, station)

    data6 = [inner_list for inner_list in data6 if len(inner_list) != 2]

    data6.insert(0, ["", "DRIVER", "DRIVER", "", "", "", "", "", "", "STATION"])


    columns = data6[0]
    columns.insert(0, "DEFAULT DATE")
    # rows = data6[1:]
    # for item in rows:
    #     print(f"length: {len(item)} | item: {item}")

    for inner_list in data6[2:]:
        if inner_list[0] != 'WEEKLY TOTALS:':
            try:
                inner_list[0] = datetime.strptime(inner_list[0], '%m/%d/%y').date()
            except ValueError as e:
                print(f"Error converting date: {e}")
    for inner_list in data6[1:]:
        inner_list.insert(0, default_date)

    for i in range(1, len(data6)):
        if len(data6[i]) != 11 and ("FEDEXID" not in data6[i]):
            data6[i][0] = str(data6[i][0])
            data6[i][1] = str(data6[i][1])
            joined_string = " <-> ".join(data6[i])
            data6[i] = [joined_string, "", "", "", "", "", "", "", "", "", ""]
            data6[i].append("Yes")
        elif "FEDEXID" in data6[i]:
            pass
        else:
            data6[i].append("No")
    rows = data6[1:]
    columns.append("Need Edit")
    df6 = pd.DataFrame(rows, columns=columns)


    # Table 7
    if "NON-SCAN ACTIVITY" in processed_details:
        table7 = processed_details.split("NON-SCAN ACTIVITY", 1)[-1]

        if "DOUBLE STOPS:" in table7:
            table7 = table7.split("DOUBLE STOPS:", 1)[0]
        else:
            table7 = table7.split("OTHER INFORMATION:", 1)[0]

        # Create DataFrame7
        data7 = [line.split() for line in table7.split('\n') if line.strip()]

        for i in range(len(data7)):
            if "WEEKLY_TOTALS:" in data7[i]:
                data7[i].insert(1, "")
                data7[i].insert(2, "")

        # rows = data7
        for inner_list in data7:
            if inner_list[0] != 'WEEKLY_TOTALS:':
                inner_list[0] = datetime.strptime(inner_list[0], '%m/%d/%y').date()

        for inner_list in data7:
            inner_list.insert(0, default_date)

        for i in range(len(data7)):
            if len(data7[i]) != 10:
                data7[i][0] = str(data7[i][0])
                data7[i][1] = str(data7[i][1])
                joined_string = " <-> ".join(data7[i])
                data7[i] = [joined_string, "", "", "", "", "", "", "", "", ""]
                data7[i].append("Yes")
            else:
                data7[i].append("No")
        rows = data7
        df7 = pd.DataFrame(rows, columns=None)
    else:
        data7 = ""
        df7 = pd.DataFrame()


    # Table 8
    if "DOUBLE STOPS:" in processed_details:
        table8 = processed_details.split("DOUBLE STOPS:", 1)[-1]
        table8 = table8.split("OTHER INFORMATION:", 1)[0]
        table8 = table8.replace("FACILITY#", "FACILITY# ")
        table8 = table8.replace(".", "")

        # Create DataFrame8
        data8 = [line.split() for line in table8.split('\n') if line.strip()]

        for j in range(1, len(data8)):
            merged_list = []
            i = 0
            while i < len(data8[j]):
                current_item = data8[j][i]
                merged_item = current_item
                while i < len(data8[j]) - 1 and current_item.isalpha() and data8[j][i + 1].isalpha(): 
                    # Continue merging consecutive alpha items with "_"
                    merged_item += "_" + data8[j][i + 1]
                    i += 1 
                merged_list.append(merged_item)
                i += 1

            data8[j] = merged_list

        columns = data8[0]
        columns.insert(0, "DEFAULT DATE")
        # rows = data8[1:]
        for inner_list in data8[1:]:
            inner_list[0] = datetime.strptime(inner_list[0], '%m/%d/%y').date()

        for inner_list in data8[1:]:
            inner_list.insert(0, default_date)

        for i in range(1, len(data8)):
            if len(data8[i]) != 5:
                data8[i][0] = str(data8[i][0])
                data8[i][1] = str(data8[i][1])
                joined_string = " <-> ".join(data8[i])
                data8[i] = [joined_string, "", "", "", ""]
                data8[i].append("Yes")
            else:
                data8[i].append("No")
        rows = data8[1:]
        columns.append("Need Edit")
        rows = data8[1:]
        df8 = pd.DataFrame(rows, columns=columns)
    else:
        data8 = ""
        df8 = pd.DataFrame(columns=['DEFAULT DATE', 'DATE', 'FACILITY#', 'DESCRIPTION', '#STOPS', "Need Edit"])


    # Table 9
    table9 = processed_details.split("OTHER INFORMATION:", 1)[-1]
    table9 = table9.replace("BASE FUEL PRICE", "BASEFUELPRICE").replace("CURRENT FUEL PRICE", "CURRENTFUELPRICE").replace("DAILY STOP THRESHOLD effective ", "DAILYSTOPTHRESHOLDeffective").replace(" :", ":").replace(" $", "")

    # Create DataFrame9
    data9 = [line.split() for line in table9.split('\n') if line.strip()]

    for i in range(len(data9)):
        data9[i].insert(0, default_date)

    for i in range(len(data9)):
        if len(data9[i]) != 3:
            data9[i][0] = str(data9[i][0])
            data9[i][1] = str(data9[i][1])
            joined_string = " <-> ".join(data9[i])
            data9[i] = [joined_string, "", "", ""]
            data9[i].append("Yes")
        else:
            data9[i].append("No")

    rows = data9
    df9 = pd.DataFrame(rows, columns=None)

    dataframe_dict["table1"] = {"data": data1, "df": df1}
    dataframe_dict["table2"] = {"data": data2, "df": df2}
    dataframe_dict["table3"] = {"data": data3, "df": df3}
    dataframe_dict["table4"] = {"data": data4, "df": df4}
    dataframe_dict["table5"] = {"data": data5, "df": df5}
    dataframe_dict["table6"] = {"data": data6, "df": df6}
    dataframe_dict["table7"] = {"data": data7, "df": df7}
    dataframe_dict["table8"] = {"data": data8, "df": df8}
    dataframe_dict["table9"] = {"data": data9, "df": df9}

    return dataframe_dict



def new_excel_file(output_excel_file, dataframe_dict, isp_id):
    # Save DataFrame to Excel with ExcelWriter
    excel_file_path = output_excel_file
    with pd.ExcelWriter(excel_file_path, engine='xlsxwriter', engine_kwargs={'options': {'strings_to_numbers': True}}) as writer:
        for i in range(1, 10):
            dataframe_dict.get(f"table{i}").get("df").to_excel(writer, index=False, sheet_name=f'Sheet{i}')

    # Load the workbook
    book = load_workbook(excel_file_path)
    writer = pd.ExcelWriter(excel_file_path, engine="openpyxl")
    writer.book = book
    writer.sheets = {ws.title: ws for ws in book.worksheets}

    # Get the sheet and insert two empty rows at the top
    sheet1 = writer.sheets['Sheet1']

    sheet1.insert_rows(1, amount=3)

    sheet1['A1'] = 'FEDEX ID'
    sheet1['B1'] = int(isp_id)

    sheet1.merge_cells('C2:F2')
    sheet1.merge_cells('G2:J2')
    sheet1.merge_cells('K2:M2')

    sheet1['C2'] = 'PICK-UP'
    sheet1['G2'] = 'DELIVERY'
    sheet1['K2'] = 'STOPS/OTHER'

    for row in sheet1['C2:K2']:
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center')

    all_border(sheet1)

    sheet1.column_dimensions['A'].width = 15
    sheet1.column_dimensions['B'].width = 15
    sheet1.column_dimensions['N'].width = 12
    sheet1.column_dimensions['O'].width = 12

    sheet1['K3'] = 'TOTAL'
    sheet1['L3'] = 'FUEL'
    sheet1['M3'] = 'SURGE STOP'

    # Sheet 2
    sheet2 = writer.sheets['Sheet2']
    all_border(sheet2)
    sheet2.column_dimensions['A'].width = 15
    sheet2.column_dimensions['B'].width = 15
    sheet2.column_dimensions['I'].width = 12

    # Sheet 3
    sheet3 = writer.sheets['Sheet3']
    all_border(sheet3)
    sheet3.column_dimensions['A'].width = 15
    sheet3.column_dimensions['B'].width = 25
    sheet3.column_dimensions['C'].width = 45
    sheet3.column_dimensions['D'].width = 12

    # Sheet 4
    sheet4 = writer.sheets['Sheet4']
    all_border(sheet4)
    sheet4.column_dimensions['A'].width = 15
    sheet4.column_dimensions['B'].width = 25
    sheet4.column_dimensions['C'].width = 20
    for col_letter in 'DEFGHIJ':
        sheet4.column_dimensions[col_letter].width = 28

    # Sheet 5
    sheet5 = writer.sheets['Sheet5']
    all_border(sheet5)
    sheet5.column_dimensions['A'].width = 15
    sheet5.column_dimensions['B'].width = 25
    sheet5.column_dimensions['C'].width = 10
    sheet5.delete_rows(1, amount=1)

    # Sheet 6
    sheet6 = writer.sheets['Sheet6']
    all_border(sheet6)
    sheet6.column_dimensions['A'].width = 20
    sheet6.column_dimensions['B'].width = 20
    sheet6.column_dimensions['C'].width = 12
    sheet6.column_dimensions['D'].width = 25
    sheet6.column_dimensions['K'].width = 27

    sheet6.merge_cells('E1:F1')
    sheet6.merge_cells('G1:H1')
    sheet6.merge_cells('I1:J1')

    sheet6['E1'] = "PICK-UP"
    sheet6['G1'] = "DELIVERY"
    sheet6['I1'] = "E-COMMERCE"

    # Sheet 7
    sheet7 = writer.sheets['Sheet7']
    all_border(sheet7)
    sheet7.column_dimensions['A'].width = 20
    sheet7.column_dimensions['B'].width = 20
    sheet7.column_dimensions['C'].width = 12
    sheet7.column_dimensions['D'].width = 12
    sheet7.delete_rows(1, amount=1)

    # Sheet 8
    sheet8 = writer.sheets['Sheet8']
    all_border(sheet8)
    sheet8.column_dimensions['A'].width = 12
    sheet8.column_dimensions['B'].width = 12
    sheet8.column_dimensions['C'].width = 12
    sheet8.column_dimensions['D'].width = 20

    # Sheet 9
    sheet9 = writer.sheets['Sheet9']
    all_border(sheet9)
    sheet9.column_dimensions['A'].width = 15
    sheet9.column_dimensions['B'].width = 40
    sheet9.delete_rows(1, amount=1)

    writer.save()



def update_excel_file(input_excel_file, dataframe_dict):
    workbook = load_workbook(input_excel_file)

    # Sheet 1
    sheet1 = workbook["Sheet1"]
    data1 = dataframe_dict.get("table1").get("data")[1:]
    data1 = to_digit(data1)
    length = len(data1[0])
    data1.insert(0, [""] * length)
    for row in data1:
        sheet1.append(row)
    all_border(sheet1)

    # Sheet 2
    sheet2 = workbook["Sheet2"]
    data2 = dataframe_dict.get("table2").get("data")[1:]
    data2 = to_digit(data2)
    length = len(data2[0])
    data2.insert(0, [""] * length)
    for row in data2:
        sheet2.append(row)
    all_border(sheet2)

    # Sheet 3
    sheet3 = workbook["Sheet3"]
    data3 = dataframe_dict.get("table3").get("data")[1:]
    data3 = to_digit(data3)
    length = len(data3[0])
    data3.insert(0, [""] * length)
    for row in data3:
        sheet3.append(row)
    all_border(sheet3)

    # Sheet 4
    if dataframe_dict.get("table4").get("data"):
        sheet4 = workbook["Sheet4"]
        last_row = sheet4.max_row
        data4 = dataframe_dict.get("table4").get("data")[1:]
        data4 = to_digit(data4)
        length = len(data4[0])
        if last_row != 1:
            data4.insert(0, [""] * length)
        for row in data4:
            sheet4.append(row)
        all_border(sheet4)

    # Sheet 5
    sheet5 = workbook["Sheet5"]
    data5 = dataframe_dict.get("table5").get("data")
    data5 = to_digit(data5)
    length = len(data5[0])
    data5.insert(0, [""] * length)
    for row in data5:
        sheet5.append(row)
    all_border(sheet5)

    # Sheet 6
    sheet6 = workbook["Sheet6"]
    data6 = dataframe_dict.get("table6").get("data")[2:]
    data6 = to_digit(data6)
    length = len(data6[0])
    data6.insert(0, [""] * length)
    for row in data6:
        sheet6.append(row)
    all_border(sheet6)

    # Sheet 7
    if dataframe_dict.get("table7").get("data"):
        sheet7 = workbook["Sheet7"]
        last_row = sheet7.max_row
        data7 = dataframe_dict.get("table7").get("data")
        data7 = to_digit(data7)
        length = len(data7[0])
        if last_row != 1:
            data7.insert(0, [""] * length)
        for row in data7:
            sheet7.append(row)
        all_border(sheet7)

    # Sheet 8
    if dataframe_dict.get("table8").get("data"):
        sheet8 = workbook["Sheet8"]
        last_row = sheet8.max_row
        data8 = dataframe_dict.get("table8").get("data")[1:]
        data8 = to_digit(data8)
        length = len(data8[0])
        if last_row != 1:
            data8.insert(0, [""] * length)
        for row in data8:
            sheet8.append(row)
        all_border(sheet8)

    # Sheet 9
    sheet9 = workbook["Sheet9"]
    data9 = dataframe_dict.get("table9").get("data")
    data9 = to_digit(data9)
    length = len(data9[0])
    data9.insert(0, [""] * length)
    for row in data9:
        sheet9.append(row)
    all_border(sheet9)

    workbook.save(input_excel_file)


class MyApp:
    def __init__(self, master):
        self.master = master
        self.master.title("Aegis Routes PDF to Excel Converter")
        self.master.iconbitmap("aegis.ico")
        self.master.geometry("500x300")
        self.master.resizable(False, False)

        self.style = ttk.Style(self.master)
        self.master.tk.call("source", "forest-dark.tcl")
        self.style.theme_use("forest-dark")

        self.file_paths = ""
        self.new_directory = ""
        self.new_file_path = ""
        self.excel_file_path = ""

        self.page1 = tk.Frame(self.master, padx=20, pady=10)
        self.page2 = tk.Frame(self.master, padx=20, pady=10)
        self.page3 = tk.Frame(self.master, padx=20, pady=10)

        self.page1.place(relx=0, rely=0, relwidth=1, relheight=1)
        self.page2.place(relx=0, rely=0, relwidth=1, relheight=1)
        self.page3.place(relx=0, rely=0, relwidth=1, relheight=1)

        # Page 1
        self.lb1 = tk.Label(self.page1, text="Select PDF file/s to convert", font=font.Font(size=12), pady=10)
        self.lb1.place(relx=0, rely=0.04, anchor="w")

        self.browse_pdf_btn = ttk.Button(self.page1, text="Browse...", command=self.choose_pdf_files, width=10)
        self.browse_pdf_btn.place(relx=0.02, rely=0.2, anchor="w")

        self.file_listbox = tk.Listbox(self.page1, selectmode=tk.MULTIPLE, width=48)
        self.file_listbox.place(relx=0.25, rely=0.45, anchor="w")

        self.btn1_1 = ttk.Button(self.page1, text="Next", command=lambda: self.go_to_page2(), state="disabled", width=10)
        self.btn1_1.place(relx=0.76, rely=0.9, anchor="w")


        # Page 2
        self.lb2 = tk.Label(self.page2, text="Choose Action", font=font.Font(size=12))
        self.lb2.place(relx=0, rely=0.04, anchor="w")

        self.option = tk.StringVar()
        self.option.set("New File")

        self.new_option = ttk.Radiobutton(self.page2, variable=self.option, text="Create New Excel File", value="New File", command=self.choose_option)
        self.new_option.place(relx=0.05, rely=0.2, anchor="w")

        self.update_option = ttk.Radiobutton(self.page2, variable=self.option, text="Update Existing Excel File", value="Update File", command=self.choose_option)
        self.update_option.place(relx=0.05, rely=0.3, anchor="w")

        self.chosen_option = self.option.get()

        self.btn2_1 = ttk.Button(self.page2, text="Back", command=lambda: self.go_to_page1(), width=10)
        self.btn2_1.place(relx=0.03, rely=0.9, anchor="w")

        self.btn2_2 = ttk.Button(self.page2, text="Next", command=lambda: self.go_to_page3(), width=10)
        self.btn2_2.place(relx=0.76, rely=0.9, anchor="w")


        # Page 3
        self.btn3_1 = ttk.Button(self.page3, text="Back", command=lambda: self.go_to_page2(), width=10)
        self.btn3_1.place(relx=0.03, rely=0.9, anchor="w")


        self.page1.tkraise()

    def choose_pdf_files(self):
        self.file_paths = filedialog.askopenfilenames(title="Select PDF file/s to convert", filetypes=[("PDF files", "*.pdf")])
        print(self.file_paths)
        if self.file_paths:
            self.btn1_1.config(state="normal")

            self.files_lb = ttk.Label(self.page1, text=f"Number of files selected: {len(self.file_paths)}")
            self.files_lb.place(relx=0.27, rely=0.8, anchor="w")

            self.file_listbox.delete(0, tk.END)
            for file_path in self.file_paths:
                self.file_listbox.insert(tk.END, basename(file_path))

    def choose_excel_file(self):
        self.excel_file_path = filedialog.askopenfilename(title="Select existing excel file to update", filetypes=[("Excel files", "*.xlsx")])
        print(self.excel_file_path)

        if self.excel_file_path:
            self.entry1.config(state=tk.NORMAL)
            self.entry1.delete(0, tk.END)
            self.entry1.insert(0, self.excel_file_path)
            self.entry1.config(state=tk.DISABLED)
        

    def choose_option(self):
        self.chosen_option = self.option.get()

    def go_to_page2(self):
        self.page2.tkraise()

    def go_to_page1(self):
        self.page1.tkraise()

    def go_to_page3(self):
        self.page3.tkraise()

        # Check if self.lb3 already exists and destroy it if it does
        if hasattr(self, 'lb3') and isinstance(self.lb3, tk.Label):
            self.lb3.destroy()

        widgets_to_destroy = ['entry', 'entry1', 'entry2', 'browse_path_btn', 'lb3_1', 'create_new_btn', 'browse_excel_btn', 'update_btn']

        for widget_name in widgets_to_destroy:
            widget = getattr(self, widget_name, None)
            if widget:
                widget.destroy()

        if self.chosen_option == "New File":
            self.lb3 = tk.Label(self.page3, text="Enter name for the output excel file", font=font.Font(size=12))
            self.lb3.place(relx=0, rely=0.41, anchor="w")

            vcmd = (self.page3.register(self.validate_filename), '%P')
            self.entry = ttk.Entry(self.page3, width=60, font=font.Font(size=10), validate='key', validatecommand=vcmd)
            self.entry.place(relx=0.03, rely=0.57, anchor="w")

            self.create_new_btn = ttk.Button(self.page3, text="Create New File", command=lambda: self.create_new_excel_file())
            self.create_new_btn.place(relx=0.71, rely=0.75, anchor="w")

            self.lb3_1 = tk.Label(self.page3, text="Select a directory to save the file", font=font.Font(size=12))
            self.lb3_1.place(relx=0, rely=0.04, anchor="w")

            self.entry2 = ttk.Entry(self.page3, width=45, font=font.Font(size=10), state=tk.DISABLED)
            self.entry2.place(relx=0.26, rely=0.2, anchor="w")

            self.browse_path_btn = ttk.Button(self.page3, text="Browse...", command=self.choose_path, width=10)
            self.browse_path_btn.place(relx=0.02, rely=0.2, anchor="w")
        else:
            self.lb3 = tk.Label(self.page3, text="Select existing excel file to update", font=font.Font(size=12))
            self.lb3.place(relx=0, rely=0.03, anchor="w")

            self.entry1 = ttk.Entry(self.page3, width=45, font=font.Font(size=10), state=tk.DISABLED)
            self.entry1.place(relx=0.26, rely=0.2, anchor="w")

            self.browse_excel_btn = ttk.Button(self.page3, text="Browse...", command=self.choose_excel_file, width=10)
            self.browse_excel_btn.place(relx=0.02, rely=0.2, anchor="w")

            self.update_btn = ttk.Button(self.page3, text="Update File", command=lambda: self.update_existing_excel_file(), width=10)
            self.update_btn.place(relx=0.765, rely=0.35, anchor="w")

    def validate_filename(self, new_value):
        # Validation function to check if the input contains forbidden characters
        forbidden_chars = ['/', '\\', ':', '*', '?', '"', '<', '>', '|']
        return all(char not in new_value for char in forbidden_chars)
    
    def choose_path(self):
        self.new_directory = filedialog.askdirectory(title="Select a directory to save the file")
        print(f"{self.new_directory}/{self.entry.get()}.xlsx")
        if self.new_directory:
            self.entry2.config(state=tk.NORMAL)
            self.entry2.delete(0, tk.END)
            self.entry2.insert(0, self.new_directory)
            self.entry2.config(state=tk.DISABLED)
        

    def create_new_excel_file(self):
        if not self.new_directory:
            messagebox.showerror("Directory missing", "Please select a directory to save the file")
        elif not self.entry.get():
            messagebox.showerror("Output filename missing", "Please enter a name for the output excel file")
        else:
            # self.new_file_path = f"{os.path.dirname(self.file_paths[0])}/{self.entry.get()}.xlsx"
            self.new_file_path = f"{self.new_directory}/{self.entry.get()}.xlsx"
            page_details = extract_pdf_data(self.file_paths[0])
            default_date, processed_details, isp_id, station, contract_num = process_text(page_details)
            print(f"--------------------------{self.file_paths[0]}")
            dataframe_dict = create_dataframe(default_date, processed_details, station, contract_num)
            new_excel_file(self.new_file_path, dataframe_dict, isp_id)

            if len(self.file_paths) > 1:
                workbook = load_workbook(self.new_file_path)
                existing_isp_id = workbook["Sheet1"]["B1"].value
                for i in range(1, len(self.file_paths)):
                    page_details = extract_pdf_data(self.file_paths[i])
                    default_date, processed_details, isp_id, station, contract_num = process_text(page_details)
                    print(f"--------------------------{self.file_paths[i]}")
                    dataframe_dict = create_dataframe(default_date, processed_details, station, contract_num)

                    if str(existing_isp_id) != str(isp_id):
                        file_name = basename(self.file_paths[i])
                        excel_file_name = basename(self.new_file_path)
                        response = messagebox.askyesnocancel("Confirm update", f"Fedex ID of {file_name} does not match the one in {excel_file_name}. Do yo wish to proceed?")
                        if response is None:  # User clicked cancel
                            proceed_flag = False
                            break
                        elif response:  # User clicked yes
                            proceed_flag = True
                            update_excel_file(self.new_file_path, dataframe_dict)

                    else:
                        proceed_flag = True
                        update_excel_file(self.new_file_path, dataframe_dict)

                if proceed_flag:
                    messagebox.showinfo("Success", "PDF converted successfully!")
                else:
                    messagebox.showinfo("Operation canceled", "Operation canceled by user.")
            else:
                messagebox.showinfo("Success", "PDF converted successfully!")

        

    def update_existing_excel_file(self):
        workbook = load_workbook(self.excel_file_path)
        existing_isp_id = workbook["Sheet1"]["B1"].value
        for file in self.file_paths:
            page_details = extract_pdf_data(file)
            default_date, processed_details, isp_id, station, contract_num = process_text(page_details)
            print(f"--------------------------{file}")
            dataframe_dict = create_dataframe(default_date, processed_details, station, contract_num)

            if str(existing_isp_id) != str(isp_id):
                file_name = basename(file)
                excel_file_name = basename(self.excel_file_path)
                response = messagebox.askyesnocancel("Confirm update", f"Fedex ID of {file_name} does not match the one in {excel_file_name}. Do yo wish to proceed?")
                if response is None:  # User clicked cancel
                    proceed_flag = False
                    break
                elif response:  # User clicked yes
                    proceed_flag = True
                    update_excel_file(self.excel_file_path, dataframe_dict)

            else:
                proceed_flag = True
                update_excel_file(self.excel_file_path, dataframe_dict)

        if proceed_flag:
            messagebox.showinfo("Success", "Excel updated successfully!")
        else:
            messagebox.showinfo("Operation canceled", "Operation canceled by user.")

  
root = tk.Tk()
app = MyApp(root)

root.mainloop()